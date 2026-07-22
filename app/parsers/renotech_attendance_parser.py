"""Renotech (Denmark) Attendance Parser
Format: Multiple sheets (one per week), decimal hours, with night shift calculation.
"""
import re
import openpyxl
from typing import List
from datetime import datetime, timedelta

class AttendanceRecord:
    def __init__(self, employee_name, date, hours, night_hours=0, role="", status="present", raw_time_slot="", subsidy_hours=0):
        self.employee_name = employee_name
        self.date = date
        self.hours = hours
        self.night_hours = night_hours
        self.subsidy_hours = subsidy_hours
        self.overtime_hours = 0
        self.role = role
        self.status = status
        self.raw_time_slot = raw_time_slot

class AttendanceSheet:
    def __init__(self, period_start="", period_end=""):
        self.period_start = period_start
        self.period_end = period_end
        self.records = []
    def add_record(self, record):
        self.records.append(record)
    def get_hours_by_employee(self, name):
        return sum(r.hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_night_hours_by_employee(self, name):
        return sum(r.night_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_total_hours(self):
        return sum(r.hours for r in self.records if r.status == "present")
    def get_total_night_hours(self):
        return sum(r.night_hours for r in self.records if r.status == "present")
    def get_total_subsidy_hours(self):
        return sum(r.subsidy_hours for r in self.records if r.status == "present")
    def get_subsidy_hours_by_employee(self, name):
        return sum(r.subsidy_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_total_overtime_hours(self):
        return sum(r.overtime_hours for r in self.records if r.status == "present")
    def get_overtime_hours_by_employee(self, name):
        return sum(r.overtime_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_employees(self):
        seen = set(); result = []
        for r in self.records:
            n = r.employee_name.strip()
            if n and n not in seen: seen.add(n); result.append(n)
        return result
    def get_records_by_employee(self, name):
        return [r for r in self.records if r.employee_name.lower().strip() == name.lower().strip()]
    def get_attendance_days(self, name):
        return sum(1 for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")

def calculate_night_hours(time_in, time_out):
    """Calculate night shift hours (18:00-06:00)."""
    NIGHT_START = 18; NIGHT_END = 30
    if time_out is None or time_in is None: return 0.0
    if time_out < time_in: time_out += 24
    overlap_start = max(time_in, NIGHT_START)
    overlap_end = min(time_out, NIGHT_END)
    return round(max(0, overlap_end - overlap_start), 2)

def parse_renotech_attendance(filepath, config=None):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = AttendanceSheet()
    for ws in wb.worksheets:
        dates = []
        for day_idx in range(7):
            col = 3 + day_idx * 3
            val = ws.cell(2, col).value
            if val:
                try:
                    dt = datetime(1899, 12, 30) + timedelta(days=int(val))
                    dates.append(dt.strftime("%Y-%m-%d"))
                except: dates.append(str(val))
            else: dates.append("")
        for r in range(5, ws.max_row + 1):
            name = ws.cell(r, 2).value
            if not name or not str(name).strip(): continue
            name = str(name).strip()
            if name.lower() == "total": continue
            for day_idx in range(7):
                tc = 3 + day_idx * 3
                tin = ws.cell(r, tc).value
                tout = ws.cell(r, tc+1).value
                sval = ws.cell(r, tc+2).value
                hrs = float(sval) if sval else 0
                if hrs <= 0: continue
                ti = float(tin) if tin is not None else None
                to = float(tout) if tout is not None else None
                night = calculate_night_hours(ti, to) if ti is not None and to is not None else 0
                raw = f"{tin}-{tout}" if tin else str(sval)
                sheet.add_record(AttendanceRecord(name, dates[day_idx] if day_idx < len(dates) else "", hrs, night, raw_time_slot=raw))
    sn = [s.strip() for s in wb.sheetnames]
    if sn:
        sheet.period_start = sn[0]
        sheet.period_end = sn[-1] if len(sn) > 1 else sn[0]
    return sheet

def parse_attendance(filepath, country="denmark", supplier="RENOTECH", config=None):
    return parse_renotech_attendance(filepath, config)
