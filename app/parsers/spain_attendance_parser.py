"""Spain Attendance Parser (Alliance / Randstad)
Format: 清关临时工记录 sheet with headers in row 1.
  A: Employee Name  B: Supplier  C: Month  D: Day
  E: Time In  F: Time Out  G: Lunch(h)  H: Hours(computed)  I: Notes
"""
import openpyxl
from datetime import time, datetime, timedelta

class AttendanceRecord:
    def __init__(self, employee_name, date, hours, night_hours=0, role="", status="present", raw_time_slot="", subsidy_hours=0):
        self.employee_name = employee_name
        self.date = date
        self.hours = hours
        self.night_hours = night_hours
        self.subsidy_hours = subsidy_hours
        self.overtime_hours = 0
        self.role = role
        self.status = status
        self.raw_time_slot = raw_time_slot

class AttendanceSheet:
    def __init__(self, period_start="", period_end=""):
        self.period_start = period_start
        self.period_end = period_end
        self.records = []
    def add_record(self, record):
        self.records.append(record)
    def get_hours_by_employee(self, name):
        return sum(r.hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_night_hours_by_employee(self, name):
        return sum(r.night_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_total_hours(self):
        return sum(r.hours for r in self.records if r.status == "present")
    def get_total_night_hours(self):
        return sum(r.night_hours for r in self.records if r.status == "present")
    def get_total_subsidy_hours(self):
        return sum(r.subsidy_hours for r in self.records if r.status == "present")
    def get_subsidy_hours_by_employee(self, name):
        return sum(r.subsidy_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_total_overtime_hours(self):
        return sum(r.overtime_hours for r in self.records if r.status == "present")
    def get_overtime_hours_by_employee(self, name):
        return sum(r.overtime_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_employees(self):
        seen = set(); result = []
        for r in self.records:
            n = r.employee_name.strip()
            if n and n not in seen: seen.add(n); result.append(n)
        return result
    def get_records_by_employee(self, name):
        return [r for r in self.records if r.employee_name.lower().strip() == name.lower().strip()]
    def get_attendance_days(self, name):
        return sum(1 for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")

def time_to_hours(t):
    """Convert time object or string to decimal hours."""
    if t is None:
        return None
    if isinstance(t, time):
        return t.hour + t.minute / 60 + t.second / 3600
    if isinstance(t, (int, float)):
        return float(t)
    try:
        parts = str(t).strip().split(":")
        return int(parts[0]) + int(parts[1]) / 60 + (int(parts[2]) / 3600 if len(parts) > 2 else 0)
    except:
        return None

def calc_night_hours(time_in_h, time_out_h):
    """Calculate night shift hours (18:00-06:00)."""
    if time_in_h is None or time_out_h is None:
        return 0.0
    NIGHT_START = 18
    NIGHT_END = 30  # 06:00 next day as 30h
    if time_out_h < time_in_h:
        time_out_h += 24
    overlap_start = max(time_in_h, NIGHT_START)
    overlap_end = min(time_out_h, NIGHT_END)
    return round(max(0, overlap_end - overlap_start), 2)

def parse_spain_attendance(filepath, supplier=None):
    """Parse Spain attendance file. If supplier is set, only parse that supplier's rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = "清关临时工记录"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    
    sheet = AttendanceSheet(period_start="", period_end="")
    year = 2026
    month_str = ""
    seen_dates = set()
    
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        
        sup_val = ws.cell(r, 2).value
        if sup_val:
            sup_val = str(sup_val).strip()
        if supplier and sup_val.upper() != supplier.upper():
            continue
        
        mon = ws.cell(r, 3).value
        if mon:
            month_str = str(mon).strip()
        
        day = ws.cell(r, 4).value
        if day is None:
            continue
        try:
            day = int(day)
        except (ValueError, TypeError):
            continue
        
        # Read hours directly from column H (already computed)
        hrs_val = ws.cell(r, 8).value
        try:
            hours = float(hrs_val) if hrs_val is not None else 0
        except (ValueError, TypeError):
            hours = 0
        
        if hours <= 0:
            continue
        
        # Time in/out for night shift calculation and raw_time
        tin = ws.cell(r, 5).value
        tout = ws.cell(r, 6).value
        lunch = ws.cell(r, 7).value
        
        tin_h = time_to_hours(tin)
        tout_h = time_to_hours(tout)
        night_h = calc_night_hours(tin_h, tout_h)
        
        # Create raw time slot string
        raw_parts = []
        if tin: raw_parts.append(str(tin))
        if tout: raw_parts.append(str(tout))
        if lunch is not None and float(lunch) > 0:
            raw_parts.append(f"lunch={lunch}h")
        raw_time = "-".join(raw_parts) if raw_parts else f"{hours}h"
        
        # Parse month to determine period
        month_map = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                     "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}
        month_num = month_map.get(month_str[:3], 6)
        
        date_str = f"{year}-{month_num:02d}-{day:02d}"
        seen_dates.add(day)
        
        sheet.add_record(AttendanceRecord(name, date_str, hours, night_hours=night_h,
                                          status="present", raw_time_slot=raw_time))
    
    # Calculate overtime for Alliance
    _calc_spain_overtime(sheet, supplier)
    
    # Determine period
    if seen_dates:
        min_day = min(seen_dates)
        max_day = max(seen_dates)
        month_num = month_map.get(month_str[:3], 6)
        sheet.period_start = f"{month_num}/{min_day}"
        sheet.period_end = f"{month_num}/{max_day}"
    
    return sheet

# Post-processing: calculate overtime for ALLIANCE only
# ALLIANCE: no subsidy_hours, no night_hours - only overtime_hours
_KNOWN_OT_SUPPLIERS = {"ALLIANCE"}

def _calc_spain_overtime(sheet, supplier):
    """Calculate overtime hours for ALLIANCE. Clear subsidy and night for non-relevant columns."""
    from collections import defaultdict
    from datetime import date
    import calendar
    
    if supplier.upper() not in {"ALLIANCE"}:
        # RANDSTAD: all zero
        for r in sheet.records:
            r.subsidy_hours = 0.0
            r.night_hours = 0.0
            r.overtime_hours = 0.0
        return
    
    # For ALLIANCE: clear everything first
    for r in sheet.records:
        r.subsidy_hours = 0.0
        r.night_hours = 0.0
        r.overtime_hours = 0.0
    
    # Group records by employee
    emp_records = defaultdict(list)
    for r in sheet.records:
        if r.status == "present":
            name = r.employee_name.strip()
            emp_records[name].append(r)
    
    # Get year/month from first record
    year, month = 2026, 6
    for recs in emp_records.values():
        for r in recs:
            try:
                parts = r.date.split("-")
                year, month = int(parts[0]), int(parts[1])
                break
            except:
                continue
        break
    
    # Calculate legal workdays x 8
    legal_days = sum(1 for d in range(1, calendar.monthrange(year, month)[1] + 1)
                    if date(year, month, d).weekday() < 5)
    threshold = legal_days * 8
    
    for name, recs in emp_records.items():
        total_hours = sum(r.hours for r in recs)
        
        # Type 2: Monthly overtime
        monthly_ot = total_hours - threshold
        if monthly_ot > 0:
            # Assign exact overtime to ALL records proportionally
            # but use round() for each record
            assigned = 0.0
            for i, r in enumerate(recs):
                if i < len(recs) - 1:
                    ot = round(r.hours / total_hours * monthly_ot, 2)
                    r.overtime_hours = ot
                    assigned += ot
                else:
                    # Last record gets the remainder for exact total
                    r.overtime_hours = round(monthly_ot - assigned, 2)
            continue
        
        # Type 1: Daily overtime
        by_date = defaultdict(list)
        for r in recs:
            by_date[r.date].append(r)
        
        daily_ot_total = 0.0
        for date_str, day_recs in by_date.items():
            day_hours = sum(r.hours for r in day_recs)
            if day_hours > 8:
                daily_ot = day_hours - 8
                # Assign proportionally across records in that day
                assigned = 0.0
                for i, r in enumerate(day_recs):
                    if i < len(day_recs) - 1:
                        ot = round(r.hours / day_hours * daily_ot, 2)
                        r.overtime_hours = ot
                        assigned += ot
                    else:
                        r.overtime_hours = round(daily_ot - assigned, 2)
                daily_ot_total += daily_ot
def parse_attendance(filepath, country="spain", supplier="ALLIANCE", config=None):
    return parse_spain_attendance(filepath, supplier)
