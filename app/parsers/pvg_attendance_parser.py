"""PVG Pacework Attendance Parser.
Format: Weekly Excel with columns per day (Mon-Sun).
  Row 2-3: Team Leader section
  Row 4-5: Section headers
  Row 6+: Regular employees
  Columns: A=name, B-D=Mon(Time,Hours,Total), E-G=Tue, H-J=Wed, K-M=Thu, N-P=Fri, Q-S=Sat, T-V=Sun, W=Weekly Total
"""
import openpyxl
import re
from datetime import datetime, time

class AttendanceRecord:
    def __init__(self, employee_name, date, hours, night_hours=0, role="", status="present", raw_time_slot="", subsidy_hours=0):
        self.employee_name = employee_name
        self.date = date
        self.hours = hours
        self.night_hours = night_hours
        self.subsidy_hours = subsidy_hours
        self.overtime_hours = 0
        self.role = role
        self.status = status
        self.raw_time_slot = raw_time_slot

class AttendanceSheet:
    def __init__(self, period_start="", period_end=""):
        self.period_start = period_start
        self.period_end = period_end
        self.records = []
    def add_record(self, record):
        self.records.append(record)
    def get_hours_by_employee(self, name):
        return sum(r.hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_night_hours_by_employee(self, name):
        return sum(r.night_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_total_hours(self):
        return sum(r.hours for r in self.records if r.status == "present")
    def get_total_night_hours(self):
        return sum(r.night_hours for r in self.records if r.status == "present")
    def get_total_subsidy_hours(self):
        return sum(r.subsidy_hours for r in self.records if r.status == "present")
    def get_subsidy_hours_by_employee(self, name):
        return sum(r.subsidy_hours for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")
    def get_total_overtime_hours(self):
        return sum(r.overtime_hours for r in self.records if r.status == "present")
    def get_employees(self):
        seen = set(); result = []
        for r in self.records:
            n = r.employee_name.strip()
            if n and n not in seen: seen.add(n); result.append(n)
        return result
    def get_records_by_employee(self, name):
        return [r for r in self.records if r.employee_name.lower().strip() == name.lower().strip()]
    def get_attendance_days(self, name):
        return sum(1 for r in self.records if r.employee_name.lower().strip() == name.lower().strip() and r.status == "present")

# Column mapping: (time_col, hours_col, total_col)
DAY_COLS = [
    (2, 3, 4),   # Monday: B, C, D
    (5, 6, 7),   # Tuesday: E, F, G
    (8, 9, 10),  # Wednesday: H, I, J
    (11, 12, 13),# Thursday: K, L, M
    (14, 15, 16),# Friday: N, O, P
    (17, 18, 19),# Saturday: Q, R, S
    (20, 21, 22),# Sunday: T, U, V
]

def parse_time_to_hours(t_val):
    """Convert a time value (string like '6:30:00' or timedelta) to decimal hours."""
    if t_val is None:
        return 0.0
    if isinstance(t_val, (int, float)):
        return float(t_val)
    s = str(t_val).strip()
    if not s:
        return 0.0
    # Handle "1 day, 18:00:00" format
    day_match = re.match(r'(?:(\d+)\s*days?,\s*)?(\d+):(\d+)(?::(\d+))?', s)
    if day_match:
        days = int(day_match.group(1)) if day_match.group(1) else 0
        hours = int(day_match.group(2))
        mins = int(day_match.group(3))
        secs = int(day_match.group(4)) if day_match.group(4) else 0
        return days * 24 + hours + mins / 60 + secs / 3600
    return 0.0

def calc_night_hours_from_time_slot(time_str, date_str=None):
    """Calculate night shift hours (18:00-06:00) from a time slot like '17:00-24:00' or '23:00-06:00'.
    Only counts night hours on weekdays (Mon-Fri)."""
    if not time_str or not isinstance(time_str, str):
        return 0.0
    # Only night hours on weekdays (Mon-Fri)
    if date_str:
        try:
            parts = date_str.split('-')
            d = date(int(parts[0]), int(parts[1]), int(parts[2]))
            if d.weekday() >= 5:  # Saturday(5) or Sunday(6)
                return 0.0
        except:
            pass
    m = re.match(r'(\d+):(\d+)\s*-\s*(\d+):(\d+)', time_str.strip())
    if not m:
        return 0.0
    start_h = int(m.group(1)) + int(m.group(2)) / 60
    end_h = int(m.group(3)) + int(m.group(4)) / 60
    if end_h < start_h:
        end_h += 24
    NIGHT_START = 18
    NIGHT_END = 30  # 06:00 next day
    overlap_start = max(start_h, NIGHT_START)
    overlap_end = min(end_h, NIGHT_END)
    return round(max(0, overlap_end - overlap_start), 2)

from collections import defaultdict
from datetime import date

def calc_subsidy_hours(sheet):
    """Calculate subsidy hours for each employee: Sat + Sun + max(0, total_week - 40)."""
    emp_records = defaultdict(list)
    for rec in sheet.records:
        if rec.status == "present":
            emp_records[rec.employee_name].append(rec)
    
    for emp_name, recs in emp_records.items():
        total_week = sum(r.hours for r in recs)
        sat_sun_hours = sum(r.hours for r in recs if _is_weekend(r.date))
        weekday_hours = sum(r.hours for r in recs if not _is_weekend(r.date))
        overtime = max(0, weekday_hours - 40)
        subsidy = round(sat_sun_hours + overtime, 2)
        for i, r in enumerate(recs):
            if i == 0:
                r.subsidy_hours = subsidy
            else:
                r.subsidy_hours = 0

def _is_weekend(date_str):
    """Check if date is Saturday(5) or Sunday(6)."""
    try:
        parts = date_str.split("-")
        d = date(int(parts[0]), int(parts[1]), int(parts[2]))
        return d.weekday() >= 5
    except:
        return False

def parse_pvg_attendance(filepath, supplier=None):
    """Parse Pacework weekly attendance Excel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    sheet = AttendanceSheet(period_start="week27", period_end="week27")
    
    # Extract dates from header row 1
    year = 2026
    day_dates = []
    # Row 1 has day headers like "MONDAY 29.06.25", "TUESDAY 30.06.25", etc.
    date_col_map = {}  # col_letter -> date_str
    for c in [2, 5, 8, 11, 14, 17, 20]:  # B, E, H, K, N, Q, T
        hdr = str(ws.cell(1, c).value or "")
        m = re.search(r'(\d+)\.(\d+)\.(\d+)', hdr)
        if m:
            d, mon, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100:
                y += 2000
            day_dates.append(f"{y}-{mon:02d}-{d:02d}")
        else:
            day_dates.append("")
    
    # Parse Team Leader (row 3)
    team_leader_name = str(ws.cell(3, 1).value or "").strip()
    if team_leader_name:
        for day_idx in range(7):
            time_col, hours_col, total_col = DAY_COLS[day_idx]
            date_str = day_dates[day_idx] if day_idx < len(day_dates) else ""
            if not date_str:
                continue
            total_val = ws.cell(3, total_col).value
            time_val = ws.cell(3, time_col).value
            hours = parse_time_to_hours(total_val)
            if hours > 0:
                night_h = calc_night_hours_from_time_slot(str(time_val or ""), date_str)
                sheet.add_record(AttendanceRecord(
                    employee_name=team_leader_name,
                    date=date_str,
                    hours=hours,
                    night_hours=night_h,
                    status="present",
                    raw_time_slot=str(time_val or ""),
                    role="Team Leader"
                ))
    
    # Parse regular employees (rows 6+)
    for r in range(6, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        
        for day_idx in range(7):
            time_col, hours_col, total_col = DAY_COLS[day_idx]
            date_str = day_dates[day_idx] if day_idx < len(day_dates) else ""
            if not date_str:
                continue
            total_val = ws.cell(r, total_col).value
            time_val = ws.cell(r, time_col).value
            hours = parse_time_to_hours(total_val)
            if hours > 0:
                night_h = calc_night_hours_from_time_slot(str(time_val or ""), date_str)
                sheet.add_record(AttendanceRecord(
                    employee_name=name,
                    date=date_str,
                    hours=hours,
                    night_hours=night_h,
                    status="present",
                    raw_time_slot=str(time_val or "")
                ))
    
    calc_subsidy_hours(sheet)
    return sheet

# Alias for generic parser interface
def parse_attendance(filepath, country="pvg", supplier="PACEWORK", config=None):
    return parse_pvg_attendance(filepath, supplier)
