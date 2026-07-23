"""Attendance parsers for Tempoteam (Belgium), Renotech (Denmark), and Spain."""
from __future__ import annotations
import sys, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from pathlib import Path
from typing import List, Optional
import openpyxl


class AttendanceRecord:
    def __init__(self, employee_name: str = "", date: str = "", hours: float = 0.0,
                 role: str = "", status: str = "present", raw_time_slot: str = "",
                 night_hours: float = 0.0, subsidy_hours: float = 0.0):
        self.employee_name = employee_name
        self.date = date
        self.hours = hours
        self.role = role
        self.status = status
        self.raw_time_slot = raw_time_slot
        self.night_hours = night_hours
        self.subsidy_hours = subsidy_hours
        self.overtime_hours = 0.0

    def __repr__(self):
        return f"{self.employee_name}: {self.date} {self.hours}h [{self.status}]"


class AttendanceSheet:
    def __init__(self, period_start: str = "", period_end: str = ""):
        self.period_start = period_start
        self.period_end = period_end
        self.records: List[AttendanceRecord] = []

    def add_record(self, rec: AttendanceRecord):
        self.records.append(rec)

    def get_records_by_employee(self, name: str) -> List[AttendanceRecord]:
        return [r for r in self.records if r.employee_name.lower().strip() == name.lower().strip()]

    def get_hours_by_employee(self, name: str) -> float:
        return sum(r.hours for r in self.records
                   if r.employee_name.lower().strip() == name.lower().strip()
                   and r.status == "present")

    def get_night_hours_by_employee(self, name: str) -> float:
        return sum(r.night_hours for r in self.records
                   if r.employee_name.lower().strip() == name.lower().strip()
                   and r.status == "present")

    def get_subsidy_hours_by_employee(self, name: str) -> float:
        vals = [r.subsidy_hours for r in self.records
                if r.employee_name.lower().strip() == name.lower().strip()
                and r.status == "present"]
        return max(vals) if vals else 0.0

    def get_attendance_days(self, name: str) -> int:
        return sum(1 for r in self.records
                   if r.employee_name.lower().strip() == name.lower().strip()
                   and r.status == "present")

    def get_total_hours(self) -> float:
        return sum(r.hours for r in self.records if r.status == "present")

    def get_total_night_hours(self) -> float:
        return sum(r.night_hours for r in self.records if r.status == "present")

    def get_total_subsidy_hours(self) -> float:
        return sum(r.subsidy_hours for r in self.records if r.status == "present")

    def get_total_overtime_hours(self) -> float:
        return sum(getattr(r, "overtime_hours", 0) for r in self.records if r.status == "present")

    def get_employees(self) -> List[str]:
        seen = set()
        result = []
        for r in self.records:
            n = r.employee_name
            if n not in seen:
                seen.add(n)
                result.append(n)
        return result


TIME_PATTERN = re.compile(r"(\d+)h(\d*)\s*[-\u2013]\s*(\d+)h(\d*)")


def detect_shift_start(time_str: str) -> int:
    """Extract the start hour from a French time slot like '6h-14h06' -> returns 6."""
    m = TIME_PATTERN.search(time_str)
    if m:
        try:
            return int(m.group(1))
        except:
            return 0
    return 0


def parse_french_time_slot(time_str: str, deduct_lunch: bool = True) -> float:
    """Parse French time format like 14h-22h26 -> decimal hours."""
    time_str = time_str.strip()
    if not time_str or time_str.upper() in ("OFF", "ABSENT", "CONGE", "MALADIE"):
        return 0.0

    m = TIME_PATTERN.search(time_str)
    if not m:
        # Try alternate format: hh:mm-hh:mm
        alt = re.search(r"(\d+):(\d+)\s*-\s*(\d+):(\d+)", time_str)
        if alt:
            sh, sm, eh, em = int(alt.group(1)), int(alt.group(2)), int(alt.group(3)), int(alt.group(4))
            start = sh + sm / 60.0
            end = eh + em / 60.0
            if end < start:
                end += 24
            raw_hours = end - start
            return max(0, raw_hours)
        return 0.0

    sh = int(m.group(1))
    sm = int(m.group(2)) if m.group(2) else 0
    eh = int(m.group(3))
    em = int(m.group(4)) if m.group(4) else 0
    start = sh + sm / 60.0
    end = eh + em / 60.0
    if end < start:
        end += 24.0
    raw_hours = end - start
    if deduct_lunch and raw_hours > 6:
        raw_hours -= 0.5
    return max(0, round(raw_hours, 2))


def extract_role_from_name(name: str) -> str:
    match = re.search(r"\(([^)]+)\)", name)
    if match:
        role_text = match.group(1)
        role_map = {
            "Conducteur de chariot \u00e9l\u00e9vateur": ["Cariste", "chariot \u00e9l\u00e9vateur", "CARISTE"],
            "Douane": ["Douane", "douane"],
            "Manoeuvre": ["Manoeuvre", "WH", "Stack", "STACK", "stack"],
            "\u00e9tudiant": ["TT", "\u00e9tudiante", "\u00e9tudiant", "Retour", "dispatch"],
        }
        for role_name, keywords in role_map.items():
            for kw in keywords:
                if kw.lower() in role_text.lower():
                    return role_name
        return role_text
    return ""


def parse_tempoteam_attendance(filepath: str, config: dict = None) -> AttendanceSheet:
    """Parse Belgium Tempoteam attendance Excel (auto-detect sheet & dates)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Auto-detect sheet
    sheet = None
    for s in wb.sheetnames:
        if "2026" in s:
            sheet = wb[s]
            break
    if sheet is None:
        sheet = wb.active

    # Detect columns with date numbers in row 2
    name_col = 1
    date_numbers = {}
    for c in range(2, sheet.max_column + 1):
        v = sheet.cell(2, c).value
        if v is not None:
            try:
                day_num = int(v)
                date_numbers[c] = day_num
            except (ValueError, TypeError):
                pass

    # Map column to day name from row 1
    fr_day_names = {"lundi": "Mon", "mardi": "Tue", "mercredi": "Wed",
                    "jeudi": "Thu", "vendredi": "Fri", "samedi": "Sat", "dimanche": "Sun"}

    att_sheet = AttendanceSheet()
    cols_with_dates = sorted(date_numbers.keys())
    if cols_with_dates:
        att_sheet.period_start = str(date_numbers[cols_with_dates[0]])
        att_sheet.period_end = str(date_numbers[cols_with_dates[-1]])

    # Determine month from sheet name
    month = 7
    sheet_lower = sheet.title.lower()
    fr_months = {"janvier": 1, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5,
                 "juin": 6, "juillet": 7, "aout": 8, "septembre": 9,
                 "octobre": 10, "novembre": 11, "decembre": 12}
    for mname, mnum in fr_months.items():
        if mname in sheet_lower:
            month = mnum
            break

    for r in range(3, sheet.max_row + 1):
        name = sheet.cell(r, name_col).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        role = extract_role_from_name(name)

        for col in cols_with_dates:
            time_val = sheet.cell(r, col).value
            if time_val is None:
                continue
            time_str = str(time_val).strip()

            upper = time_str.upper()
            if upper == "OFF":
                status = "off"
                hours = 0.0
            elif upper == "ABSENT":
                status = "absent"
                hours = 0.0
            else:
                status = "present"
                hours = parse_french_time_slot(time_str, deduct_lunch=True)

            subsidy_h = 0.0
            if status == "present" and hours > 0:
                start_h = detect_shift_start(time_str)
                subsidy_h = hours if start_h in (6, 11, 14) else 0.0

            day = date_numbers.get(col, 0)
            if day > 0:
                date_str = f"2026-{month:02d}-{day:02d}"
                att_sheet.add_record(AttendanceRecord(
                    name, date_str, hours, role, status, time_str,
                    night_hours=0.0, subsidy_hours=subsidy_h
                ))

    return att_sheet


from .renotech_attendance_parser import parse_renotech_attendance
from .spain_attendance_parser import parse_attendance as parse_spain_attendance
from .pvg_attendance_parser import parse_attendance as parse_pvg_attendance
from .pvg_worksupply_attendance_parser import parse_attendance as parse_worksupply_attendance
from .pvg_attendance_parser import parse_attendance as parse_pvg_attendance
from .pvg_worksupply_attendance_parser import parse_attendance as parse_worksupply_attendance


def parse_attendance(filepath: str, country: str = "belgium",
                     supplier: str = "TEMPOTEAM", config: dict = None) -> AttendanceSheet:
    """Route to the correct parser based on country and supplier."""
    country_lower = country.lower().strip()
    supplier_upper = supplier.upper().strip()
    if country_lower == "denmark" and supplier_upper == "RENOTECH":
        return parse_renotech_attendance(filepath, config)
    if country_lower == "spain":
        return parse_spain_attendance(filepath, supplier=supplier)
    if country_lower == "pvg":
        if supplier_upper == "WORKSUPPLY":
            return parse_worksupply_attendance(filepath, supplier=supplier)
        return parse_pvg_attendance(filepath, supplier=supplier)
    return parse_tempoteam_attendance(filepath, config)


if __name__ == "__main__":
    pass

