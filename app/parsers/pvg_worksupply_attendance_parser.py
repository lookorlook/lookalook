"""PVG Worksupply Attendance Parser.
Format similar to Pacework but with different column layout.
  Row 2-4: Team Leader section (Team Leader header, then employee rows)
  Row 5: Team Leader subtotals
  Row 6-7: Section headers for regular employees
  Row 8+: Regular employees
  Columns: A=name, B-D=Mon(Time,Hours,Total), E-G=Tue, H-J=Wed, K-M=Thu, N-P=Fri, Q-S=Sat, T-U=Sun(Time,Hours)
  Sunday also has a Total column (V)
"""
import sys, os, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from datetime import datetime, time

# Reuse the same base classes
from parsers.pvg_attendance_parser import AttendanceRecord, AttendanceSheet, parse_time_to_hours, calc_subsidy_hours

def calc_night_hours_ws(time_str, date_str=None):
    """Calculate night shift hours (19:00-06:00) for Worksupply.
    Only counts night hours on weekdays (Mon-Fri)."""
    if not time_str or not isinstance(time_str, str):
        return 0.0
    # Only night hours on weekdays (Mon-Fri)
    if date_str:
        try:
            parts = date_str.split('-')
            d = __import__('datetime').date(int(parts[0]), int(parts[1]), int(parts[2]))
            if d.weekday() >= 5:  # Saturday(5) or Sunday(6)
                return 0.0
        except:
            pass
    m = re.match(r"(\d+)[:;](\d+)\s*-(\d+)[:;](\d+)", time_str.strip())
    if not m:
        return 0.0
    start_h = int(m.group(1)) + int(m.group(2)) / 60
    end_h = int(m.group(3)) + int(m.group(4)) / 60
    if end_h < start_h:
        end_h += 24
    NIGHT_START = 19
    NIGHT_END = 30  # 06:00 next day
    overlap_start = max(start_h, NIGHT_START)
    overlap_end = min(end_h, NIGHT_END)
    return round(max(0, overlap_end - overlap_start), 2)

# Column mapping for Worksupply: (time_col, hours_col, total_col)
# All days have Time, Hours, Total columns
DAY_COLS_WS = [
    (2, 3, 4),   # Monday: B, C, D
    (5, 6, 7),   # Tuesday: E, F, G
    (8, 9, 10),  # Wednesday: H, I, J
    (11, 12, 13),# Thursday: K, L, M
    (14, 15, 16),# Friday: N, O, P
    (17, 18, 19),# Saturday: Q, R, S
    (20, 21, 22),# Sunday: T, U, V
]

def parse_worksupply_attendance(filepath, supplier=None):
    """Parse Worksupply weekly attendance Excel."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    sheet = AttendanceSheet(period_start="week27", period_end="week27")
    
    # Extract dates from header row 1
    day_dates = []
    for c in [2, 5, 8, 11, 14, 17, 20]:  # B, E, H, K, N, Q, T
        hdr = str(ws.cell(1, c).value or "")
        m = re.search(r"(\d+)\.(\d+)\.(\d+)", hdr)
        if m:
            d, mon, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100:
                y += 2000
            if y < 2026:
                y = 2026
            day_dates.append(f"{y}-{mon:02d}-{d:02d}")
        else:
            day_dates.append("")
    
    # Parse Team Leaders (rows 3-4)
    # Row 2 is "Team Leader" header, row 3+ are employee rows
    for r in range(3, 5):  # Rows 3-4 (Stefin Saji, Jaspreet Singh)
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        
        for day_idx in range(7):
            time_col, hours_col, total_col = DAY_COLS_WS[day_idx]
            date_str = day_dates[day_idx] if day_idx < len(day_dates) else ""
            if not date_str:
                continue
            
            time_val = ws.cell(r, time_col).value
            time_str = str(time_val or "").strip() if time_val else ""
            
            total_val = ws.cell(r, total_col).value
            hours = parse_time_to_hours(total_val) if total_val else 0.0
            
            if hours > 0:
                night_h = calc_night_hours_ws(time_str, date_str)
                sheet.add_record(AttendanceRecord(
                    employee_name=name,
                    date=date_str,
                    hours=hours,
                    night_hours=night_h,
                    status="present",
                    raw_time_slot=time_str,
                    role="Team Leader"
                ))
    
    # Parse regular employees (dynamic start: skip header rows, employees start after subtotal)
    emp_start = 6
    for cr in range(6, min(10, ws.max_row + 1)):
        cv = str(ws.cell(cr, 1).value or "").strip()
        if cv in ("", "Work Supply", "Employee name"):
            emp_start = cr + 1
        else:
            # Found first employee name
            emp_start = cr
            break
    
    for r in range(emp_start, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        
        for day_idx in range(7):
            time_col, hours_col, total_col = DAY_COLS_WS[day_idx]
            date_str = day_dates[day_idx] if day_idx < len(day_dates) else ""
            if not date_str:
                continue
            
            time_val = ws.cell(r, time_col).value
            time_str = str(time_val or "").strip() if time_val else ""
            
            # Skip OFF and sick
            if time_str.upper() in ("OFF", "SICK", ""):
                continue
            
            total_val = ws.cell(r, total_col).value
            hours = parse_time_to_hours(total_val) if total_val else 0.0
            
            if hours > 0:
                night_h = calc_night_hours_ws(time_str, date_str)
                sheet.add_record(AttendanceRecord(
                    employee_name=name,
                    date=date_str,
                    hours=hours,
                    night_hours=night_h,
                    status="present",
                    raw_time_slot=time_str
                ))
    
    calc_subsidy_hours(sheet)
    return sheet

def parse_attendance(filepath, country="pvg", supplier="WORKSUPPLY", config=None):
    return parse_worksupply_attendance(filepath, supplier)
